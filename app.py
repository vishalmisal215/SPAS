from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, flash, jsonify, send_file
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, random, json, io, logging, traceback

app = Flask(__name__)
app.secret_key = "super-secret-localhost-key-2026"
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Configure logging
logging.basicConfig(
    level=logging.ERROR,
    format='%(asctime)s %(levelname)s %(name)s %(message)s'
)
logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
USERS_FILE = os.path.join(DATA_DIR, "users.json")
FACULTY_FILE = os.path.join(DATA_DIR, "faculty.json")
PRACTICALS_FILE = os.path.join(DATA_DIR, "practicals.json")
QUESTIONS_FILE = os.path.join(DATA_DIR, "questions.json")
SUBJECTS_FILE = os.path.join(DATA_DIR, "subjects.json")
RESULTS_DIR = os.path.join(DATA_DIR, "results")

EXAM_DURATION_SECONDS = 30 * 60

# ─────────────────────────── Global Error Handlers ────────────────────────────

@app.errorhandler(404)
def not_found_error(error):
    flash("Page not found.", "error")
    if is_faculty():
        return redirect(url_for("faculty_dashboard"))
    if is_student():
        return redirect(url_for("dashboard"))
    return redirect(url_for("index"))

@app.errorhandler(500)
def internal_error(error):
    logger.error("Internal Server Error: %s\n%s", error, traceback.format_exc())
    flash("An internal error occurred. Please try again.", "error")
    if is_faculty():
        return redirect(url_for("faculty_dashboard"))
    if is_student():
        return redirect(url_for("dashboard"))
    return redirect(url_for("index"))

@app.errorhandler(Exception)
def unhandled_exception(e):
    logger.error("Unhandled Exception: %s\n%s", e, traceback.format_exc())
    flash("An unexpected error occurred. Please try again.", "error")
    if is_faculty():
        return redirect(url_for("faculty_dashboard"))
    if is_student():
        return redirect(url_for("dashboard"))
    return redirect(url_for("index"))

# ─────────────────────────── Helpers ──────────────────────────────────────────

def read_file_safe(filepath):
    """Read a file trying utf-8 first, then latin-1 as fallback. Returns content string or None."""
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            with open(filepath, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except Exception as e:
            logger.error("Error reading file %s: %s", filepath, e)
            return None
    return None

def load_json(path):
    if not os.path.exists(path):
        if "users" in path or "faculty" in path:
            return {}
        return []
    try:
        content = read_file_safe(path)
        if not content or not content.strip():
            return {} if ("users" in path or "faculty" in path) else []
        return json.loads(content)
    except Exception as e:
        logger.error("Could not parse %s: %s", path, e)
        return {} if ("users" in path or "faculty" in path) else []

def save_json(path, data):
    try:
        dir_path = os.path.dirname(path)
        if dir_path:
            os.makedirs(dir_path, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        logger.error("Could not save %s: %s", path, e)
        return False

def is_student():
    return "roll_no" in session

def is_faculty():
    return "faculty_id" in session

def is_logged_in():
    return is_student() or is_faculty()

def delete_user_results(roll_no):
    """Delete all result files for a specific user."""
    if not os.path.exists(RESULTS_DIR):
        return
    for filename in os.listdir(RESULTS_DIR):
        if filename.startswith(f"Result_RollNo_{roll_no}_") and filename.endswith('.txt'):
            filepath = os.path.join(RESULTS_DIR, filename)
            try:
                os.remove(filepath)
            except Exception as e:
                logger.error("Could not delete %s: %s", filepath, e)

def get_student_results(roll_no):
    results = []
    if not os.path.exists(RESULTS_DIR):
        return results
    for filename in os.listdir(RESULTS_DIR):
        if not (filename.startswith(f"Result_RollNo_{roll_no}_") and filename.endswith('.txt')):
            continue
        filepath = os.path.join(RESULTS_DIR, filename)
        content = read_file_safe(filepath)
        if content is None:
            continue
        result_dict = {}
        for line in content.split('\n'):
            if '==========' in line:
                break
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip()
                if key and key not in ('----------',):
                    result_dict[key] = value.strip()
        results.append(result_dict)
    return results

def get_practical_questions(practical_name):
    questions = load_json(QUESTIONS_FILE)
    return [q for q in questions if q.get('practical') == practical_name]

def get_questions_by_ids(id_list):
    """Fetch full question objects matching a list of IDs (preserves order)."""
    try:
        questions = load_json(QUESTIONS_FILE)
        id_set = {int(i) for i in id_list}
        by_id = {int(q["id"]): q for q in questions if int(q.get("id", -1)) in id_set}
        return [by_id[int(i)] for i in id_list if int(i) in by_id]
    except Exception as e:
        logger.error("get_questions_by_ids error: %s", e)
        return []

def get_subject_practicals(subject_id):
    """Get all practicals for a specific subject."""
    subjects = load_json(SUBJECTS_FILE)
    for subject in subjects:
        if subject.get('id') == subject_id:
            return subject.get('practicals', [])
    return []

def get_all_practicals_for_subject(subject_name):
    """Get practicals by subject name."""
    subjects = load_json(SUBJECTS_FILE)
    for subject in subjects:
        if subject.get('name') == subject_name:
            return subject.get('practicals', [])
    return []

def extract_practical_number(name):
    """Extract the practical number from names like 'Practical No: 1 ...'"""
    import re
    match = re.search(r'(?i)practical\s*(?:no\.?\s*)?:?\s*(\d+)', name)
    if match:
        return int(match.group(1))
    match = re.search(r'^\s*(\d+)', name)
    if match:
        return int(match.group(1))
    return 9999

def insert_practical_sorted(practicals_list, new_practical):
    """Insert new_practical into the list in ascending practical-number order."""
    new_num = extract_practical_number(new_practical)
    for i, p in enumerate(practicals_list):
        if extract_practical_number(p) > new_num:
            practicals_list.insert(i, new_practical)
            return practicals_list
    practicals_list.append(new_practical)
    return practicals_list

def parse_result_file(filepath):
    """Parse result file and extract detailed question-wise results. Returns dict."""
    try:
        content = read_file_safe(filepath)
        if not content:
            return {}

        result_dict = {}
        lines = content.split('\n')

        for line in lines:
            if '==========' in line:
                break
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip()
                if key:
                    result_dict[key] = value.strip()

        # Extract detailed answers
        detailed_answers = []
        in_questions_section = False
        current_question = {}

        for line in lines:
            if "QUESTION WISE RESULT" in line:
                in_questions_section = True
                continue

            if in_questions_section:
                line = line.strip()

                if line.startswith('Q') and '. ' in line:
                    if current_question:
                        detailed_answers.append(current_question)
                    current_question = {
                        'question': line.split('. ', 1)[1] if '. ' in line else '',
                        'options': []
                    }
                elif line.startswith(('A) ', 'B) ', 'C) ', 'D) ')):
                    current_question['options'].append(line)
                elif line.startswith('Your Answer'):
                    current_question['student_answer'] = line.split(':', 1)[1].strip() if ':' in line else ''
                elif line.startswith('Correct Answer'):
                    current_question['correct_answer'] = line.split(':', 1)[1].strip() if ':' in line else ''
                elif line.startswith('Status'):
                    current_question['status'] = line.split(':', 1)[1].strip() if ':' in line else ''

        if current_question:
            detailed_answers.append(current_question)

        result_dict['detailed_answers'] = detailed_answers
        return result_dict

    except Exception as e:
        logger.error("parse_result_file error for %s: %s", filepath, e)
        return {}

def _find_result_file(roll_no, practical_name):
    """Find the latest result file for a roll_no + practical_name combo. Returns (filepath, filename) or (None, None)."""
    if not os.path.exists(RESULTS_DIR):
        return None, None
    best_ts = -1
    best_path = None
    best_name = None
    for filename in os.listdir(RESULTS_DIR):
        if not (filename.startswith(f"Result_RollNo_{roll_no}_") and filename.endswith('.txt')):
            continue
        filepath = os.path.join(RESULTS_DIR, filename)
        content = read_file_safe(filepath)
        if content and f"Practical: {practical_name}" in content:
            try:
                ts = int(filename.replace('.txt', '').rsplit('_', 1)[-1])
            except Exception:
                ts = 0
            if ts > best_ts:
                best_ts = ts
                best_path = filepath
                best_name = filename
    return best_path, best_name

# ─────────────────────────── Routes ───────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def index():
    try:
        users = load_json(USERS_FILE)
        faculty_data = load_json(FACULTY_FILE)

        if request.method == "POST":
            login_type = request.form.get("login_type")
            action = request.form.get("action")

            if login_type == "student":
                if action == "login":
                    roll = request.form.get("roll_no", "").strip()
                    password = request.form.get("password", "").strip()
                    user = users.get(roll)

                    if user and user.get("password") == password:
                        session["roll_no"] = roll
                        session["full_name"] = user.get("full_name", "")
                        session["branch"] = user.get("branch", "")
                        session["year"] = user.get("year", "")
                        session["batch"] = user.get("batch", "1")
                        session["email"] = user.get("email", "")
                        session["user_type"] = "student"
                        flash("Login successful.", "success")
                        return redirect(url_for("dashboard"))
                    else:
                        flash("Invalid roll number or password.", "error")

                elif action == "register":
                    roll = request.form.get("roll_no", "").strip()
                    password = request.form.get("password", "").strip()
                    full_name = request.form.get("full_name", "").strip()
                    branch = request.form.get("branch", "")
                    year = request.form.get("year", "")
                    batch = request.form.get("batch", "1")
                    email = request.form.get("email", "").strip()

                    if not roll or not password or len(password) < 6:
                        flash("Roll no + password required, password must be at least 6 characters.", "error")
                    elif roll in users:
                        flash("Roll number already exists.", "error")
                    else:
                        users[roll] = {
                            "roll_no": roll,
                            "password": password,
                            "full_name": full_name,
                            "branch": branch,
                            "year": year,
                            "batch": batch,
                            "email": email
                        }
                        save_json(USERS_FILE, users)
                        flash("Profile created successfully. Now you can login.", "success")

            elif login_type == "faculty":
                if action == "login":
                    faculty_id = request.form.get("faculty_id", "").strip()
                    password = request.form.get("password", "").strip()
                    faculty = faculty_data.get(faculty_id)

                    if faculty and faculty.get("password") == password:
                        session["faculty_id"] = faculty_id
                        session["full_name"] = faculty.get("full_name", "")
                        session["department"] = faculty.get("department", "")
                        session["email"] = faculty.get("email", "")
                        session["user_type"] = "faculty"
                        flash("Login successful.", "success")
                        return redirect(url_for("faculty_dashboard"))
                    else:
                        flash("Invalid faculty ID or password.", "error")

                elif action == "register":
                    faculty_id = request.form.get("faculty_id", "").strip()
                    password = request.form.get("password", "").strip()
                    full_name = request.form.get("full_name", "").strip()
                    department = request.form.get("department", "")
                    email = request.form.get("email", "").strip()

                    if not faculty_id or not password or len(password) < 6:
                        flash("Faculty ID + password required, password must be at least 6 characters.", "error")
                    elif faculty_id in faculty_data:
                        flash("Faculty ID already exists.", "error")
                    else:
                        faculty_data[faculty_id] = {
                            "faculty_id": faculty_id,
                            "password": password,
                            "full_name": full_name,
                            "department": department,
                            "email": email
                        }
                        save_json(FACULTY_FILE, faculty_data)
                        flash("Faculty profile created successfully. Now you can login.", "success")

        return render_template("login.html")
    except Exception as e:
        logger.error("index error: %s\n%s", e, traceback.format_exc())
        flash("An error occurred. Please try again.", "error")
        return render_template("login.html")


@app.route("/forgot_password", methods=["POST"])
def forgot_password():
    return jsonify({
        "success": False,
        "message": "Password recovery is disabled. Please contact administrator."
    })


@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    if not is_student():
        return redirect(url_for("index"))

    try:
        users = load_json(USERS_FILE)
        user = users.get(session["roll_no"])
        if not user:
            session.clear()
            flash("User not found. Please login again.", "error")
            return redirect(url_for("index"))

        subjects = load_json(SUBJECTS_FILE)
        first_subject = subjects[0]['name'] if subjects else 'all'
        session_subject = session.get('selected_subject', first_subject)
        selected_subject = request.args.get('subject', session_subject)
        valid_names = [s['name'] for s in subjects]
        if selected_subject not in valid_names and selected_subject != 'all':
            selected_subject = first_subject
        session['selected_subject'] = selected_subject

        if selected_subject == 'all':
            practicals = load_json(PRACTICALS_FILE)
        else:
            practicals = get_all_practicals_for_subject(selected_subject)

        student_results = get_student_results(session["roll_no"])
        submitted_practicals = {r.get("Practical", "") for r in student_results}

        if request.method == "POST":
            practical_name = request.form.get("practical_name", "").strip()

            if not practical_name:
                flash("Invalid practical selected.", "error")
                return redirect(url_for("dashboard", subject=selected_subject))

            if practical_name in submitted_practicals:
                flash("You have already submitted this practical.", "error")
                return redirect(url_for("dashboard", subject=selected_subject))

            questions = get_practical_questions(practical_name)

            if not questions:
                flash("No questions available for this practical. Please contact faculty.", "error")
                return redirect(url_for("dashboard", subject=selected_subject))

            random.shuffle(questions)
            selected = questions[:20]

            for key in ["last_result", "last_result_file", "exam_question_ids",
                        "exam_questions", "exam_start_time", "exam_duration", "practical_name"]:
                session.pop(key, None)

            session["exam_question_ids"] = [int(q["id"]) for q in selected]
            session["exam_start_time"] = datetime.now().timestamp()
            session["exam_duration"] = EXAM_DURATION_SECONDS
            session["practical_name"] = practical_name
            session.modified = True
            return redirect(url_for("exam"))

        return render_template("dashboard.html", title="Dashboard", user=user, practicals=practicals,
                               submitted_practicals=submitted_practicals, student_results=student_results,
                               subjects=subjects, selected_subject=selected_subject)
    except Exception as e:
        logger.error("dashboard error: %s\n%s", e, traceback.format_exc())
        flash("An error occurred loading the dashboard.", "error")
        return redirect(url_for("index"))


@app.route("/faculty_dashboard")
def faculty_dashboard():
    if not is_faculty():
        return redirect(url_for("index"))

    try:
        faculty_data = load_json(FACULTY_FILE)
        faculty = faculty_data.get(session["faculty_id"])
        all_students = load_json(USERS_FILE)
        practicals = load_json(PRACTICALS_FILE)
        subjects = load_json(SUBJECTS_FILE)

        selected_batch = request.args.get('batch', 'all')
        first_subject = subjects[0]['name'] if subjects else 'all'
        selected_subject = request.args.get('subject', first_subject)

        if selected_batch != 'all':
            students = {k: v for k, v in all_students.items() if v.get('batch', '1') == selected_batch}
        else:
            students = all_students

        if selected_subject != 'all':
            practicals = get_all_practicals_for_subject(selected_subject)

        all_batches = sorted(set(s.get('batch', '1') for s in all_students.values()))

        # Collect all results (latest per student+practical)
        _raw_results = {}
        if os.path.exists(RESULTS_DIR):
            for filename in os.listdir(RESULTS_DIR):
                if not filename.endswith('.txt'):
                    continue
                filepath = os.path.join(RESULTS_DIR, filename)
                raw = read_file_safe(filepath)
                if not raw:
                    continue
                result_dict = {}
                for line in raw.split('\n'):
                    if '==========' in line:
                        break
                    if ':' in line:
                        _k, _v = line.split(':', 1)
                        _k = _k.strip()
                        if _k:
                            result_dict[_k] = _v.strip()
                practical_nm = result_dict.get('Practical', '').strip()
                roll_nm = result_dict.get('Roll No', '').strip()
                if not practical_nm or not roll_nm:
                    continue
                try:
                    ts = int(filename.replace('.txt', '').rsplit('_', 1)[-1])
                except Exception:
                    ts = 0
                dedup_key = (roll_nm, practical_nm)
                if dedup_key not in _raw_results or ts > _raw_results[dedup_key][0]:
                    _raw_results[dedup_key] = (ts, result_dict)

        all_results = [v for _, v in _raw_results.values()]

        # Student performance
        student_performance = {}
        for student_id, student_data in students.items():
            student_results = [r for r in all_results if r.get("Roll No") == student_id]
            practical_scores = {}
            total_score = 0
            count = 0

            for result in student_results:
                practical_name = result.get("Practical", "").strip()
                if not practical_name:
                    continue
                practicals_stripped = [p.strip() for p in practicals]
                if selected_subject != 'all' and practical_name not in practicals_stripped:
                    continue
                try:
                    canonical = practicals[practicals_stripped.index(practical_name)]
                except (ValueError, IndexError):
                    canonical = practical_name

                score_str = result.get("Score", "0 / 0")
                if '/' in score_str:
                    try:
                        score = int(score_str.split('/')[0].strip())
                    except (ValueError, TypeError):
                        score = 0
                    if canonical not in practical_scores:
                        practical_scores[canonical] = score
                        total_score += score
                        count += 1

            avg_score = round(total_score / count, 2) if count > 0 else 0
            student_performance[student_id] = {
                "name": student_data.get("full_name", ""),
                "branch": student_data.get("branch", ""),
                "year": student_data.get("year", ""),
                "batch": student_data.get("batch", "1"),
                "email": student_data.get("email", ""),
                "practical_scores": practical_scores,
                "total": total_score,
                "average": avg_score,
                "exams_taken": count
            }

        # Practical submissions
        practical_submissions = {}
        for practical_name in practicals:
            submitted_students = []
            pname_stripped = practical_name.strip()
            seen_rolls = set()
            for result in all_results:
                if result.get("Practical", "").strip() == pname_stripped:
                    roll_no = result.get("Roll No")
                    if roll_no and roll_no in students and roll_no not in seen_rolls:
                        student = students.get(roll_no)
                        if student:
                            submitted_students.append({
                                'roll_no': roll_no,
                                'name': student.get('full_name', ''),
                                'batch': student.get('batch', '1')
                            })
                            seen_rolls.add(roll_no)
            practical_submissions[practical_name] = submitted_students

        return render_template("faculty_dashboard.html", title="Faculty Dashboard", faculty=faculty,
                               students=students, all_students=all_students, practicals=practicals,
                               results=all_results, student_performance=student_performance,
                               practical_submissions=practical_submissions, all_batches=all_batches,
                               selected_batch=selected_batch, subjects=subjects, selected_subject=selected_subject)
    except Exception as e:
        logger.error("faculty_dashboard error: %s\n%s", e, traceback.format_exc())
        flash("An error occurred loading the faculty dashboard.", "error")
        return redirect(url_for("index"))


@app.route("/api/add_subject", methods=["POST"])
def add_subject():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json(silent=True) or {}
        subject_name = data.get("name", "").strip()

        if not subject_name:
            return jsonify({"success": False, "message": "Subject name required"}), 400

        subjects = load_json(SUBJECTS_FILE)

        for subject in subjects:
            if subject.get('name', '').lower() == subject_name.lower():
                return jsonify({"success": False, "message": "Subject already exists"}), 400

        # Safely generate new ID
        try:
            max_id = max(int(s['id']) for s in subjects if str(s.get('id', '')).isdigit())
            new_id = str(max_id + 1)
        except (ValueError, TypeError):
            new_id = str(len(subjects) + 1)

        subjects.append({"id": new_id, "name": subject_name, "practicals": []})
        save_json(SUBJECTS_FILE, subjects)

        return jsonify({"success": True, "subject": {"id": new_id, "name": subject_name}}), 200
    except Exception as e:
        logger.error("add_subject error: %s", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/api/add_practical", methods=["POST"])
def add_practical():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json(silent=True) or {}
        practical_name = data.get("name", "").strip()
        subject_id = str(data.get("subject_id", "1"))

        if not practical_name:
            return jsonify({"success": False, "message": "Practical name required"}), 400

        practicals = load_json(PRACTICALS_FILE)

        if practical_name in practicals:
            return jsonify({"success": False, "message": "Practical already exists"}), 400

        practicals = insert_practical_sorted(practicals, practical_name)
        save_json(PRACTICALS_FILE, practicals)

        subjects = load_json(SUBJECTS_FILE)
        for subject in subjects:
            if str(subject.get('id')) == subject_id:
                if practical_name not in subject['practicals']:
                    subject['practicals'] = insert_practical_sorted(subject['practicals'], practical_name)
                break
        save_json(SUBJECTS_FILE, subjects)

        return jsonify({"success": True, "practical": practical_name}), 200
    except Exception as e:
        logger.error("add_practical error: %s", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/api/remove_practical", methods=["POST"])
def remove_practical():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json(silent=True) or {}
        practical_name = data.get("name", "").strip()

        practicals = load_json(PRACTICALS_FILE)

        if practical_name not in practicals:
            return jsonify({"success": False, "message": "Practical not found"}), 404

        practicals.remove(practical_name)
        save_json(PRACTICALS_FILE, practicals)

        subjects = load_json(SUBJECTS_FILE)
        for subject in subjects:
            if practical_name in subject.get('practicals', []):
                subject['practicals'].remove(practical_name)
        save_json(SUBJECTS_FILE, subjects)

        return jsonify({"success": True}), 200
    except Exception as e:
        logger.error("remove_practical error: %s", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/export_excel")
def export_excel():
    if not is_faculty():
        return redirect(url_for("index"))

    try:
        all_students = load_json(USERS_FILE)
        practicals = load_json(PRACTICALS_FILE)
        selected_subject = request.args.get('subject', 'all')

        if selected_subject != 'all':
            practicals = get_all_practicals_for_subject(selected_subject)

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Performance"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center")
        bold_font = Font(bold=True)

        headers = ["Roll No", "Name", "Branch", "Year", "Batch"] + list(practicals) + ["Total", "Average"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        row = 2
        for roll_no, student in all_students.items():
            # Basic info
            for col_idx, val in enumerate([
                roll_no,
                student.get("full_name", ""),
                student.get("branch", ""),
                student.get("year", ""),
                student.get("batch", "1")
            ], 1):
                ws.cell(row=row, column=col_idx, value=val).border = border

            student_results = get_student_results(roll_no)
            scores_dict = {}
            for result in student_results:
                pname = result.get("Practical")
                if pname and pname in practicals:
                    score_str = result.get("Score", "0 / 0")
                    if '/' in score_str:
                        try:
                            scores_dict[pname] = int(score_str.split('/')[0].strip())
                        except (ValueError, TypeError):
                            scores_dict[pname] = 0

            col = 6
            total = 0
            count = 0
            for practical in practicals:
                score = scores_dict.get(practical)
                if score is not None:
                    cell = ws.cell(row=row, column=col, value=score)
                    cell.alignment = center_align
                    total += score
                    count += 1
                else:
                    cell = ws.cell(row=row, column=col, value="-")
                cell.border = border
                col += 1

            # Total
            total_cell = ws.cell(row=row, column=col, value=total)
            total_cell.border = border
            total_cell.alignment = center_align
            total_cell.font = bold_font

            # Average
            avg = round(total / count, 2) if count > 0 else 0
            avg_cell = ws.cell(row=row, column=col + 1, value=avg)
            avg_cell.border = border
            avg_cell.alignment = center_align
            avg_cell.font = bold_font

            row += 1

        # Auto column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        subject_suffix = f"_{selected_subject}" if selected_subject != 'all' else ""
        filename = f"student_performance{subject_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error("export_excel error: %s\n%s", e, traceback.format_exc())
        flash(f"Error exporting to Excel: {str(e)}", "error")
        return redirect(url_for("faculty_dashboard"))


@app.route("/delete_account", methods=["POST"])
def delete_account():
    try:
        if is_student():
            roll_no = session["roll_no"]
            users = load_json(USERS_FILE)
            if roll_no in users:
                del users[roll_no]
                save_json(USERS_FILE, users)
            delete_user_results(roll_no)
            session.clear()
            flash("Your account has been deleted successfully.", "success")
            return redirect(url_for("index"))

        elif is_faculty():
            faculty_id = session["faculty_id"]
            faculty = load_json(FACULTY_FILE)
            if faculty_id in faculty:
                del faculty[faculty_id]
                save_json(FACULTY_FILE, faculty)
            session.clear()
            flash("Your account has been deleted successfully.", "success")
            return redirect(url_for("index"))

    except Exception as e:
        logger.error("delete_account error: %s", e)
        flash("Error deleting account.", "error")

    return redirect(url_for("index"))


@app.route("/update_profile", methods=["POST"])
def update_profile():
    if not is_student():
        return redirect(url_for("index"))
    try:
        users = load_json(USERS_FILE)
        roll_no = session["roll_no"]

        full_name = request.form.get("full_name", "").strip()
        branch = request.form.get("branch", "")
        year = request.form.get("year", "")
        batch = request.form.get("batch", "1")
        email = request.form.get("email", "").strip()

        if roll_no in users:
            users[roll_no].update({
                "full_name": full_name, "branch": branch,
                "year": year, "batch": batch, "email": email
            })
            session.update({
                "full_name": full_name, "branch": branch,
                "year": year, "batch": batch, "email": email
            })
            save_json(USERS_FILE, users)
            flash("Profile updated successfully!", "success")
    except Exception as e:
        logger.error("update_profile error: %s", e)
        flash("Error updating profile.", "error")

    return redirect(url_for("dashboard"))


@app.route("/faculty/update_profile", methods=["POST"])
def faculty_update_profile():
    if not is_faculty():
        return redirect(url_for("index"))
    try:
        faculty_data = load_json(FACULTY_FILE)
        faculty_id = session["faculty_id"]

        full_name = request.form.get("full_name", "").strip()
        department = request.form.get("department", "")
        email = request.form.get("email", "").strip()

        if faculty_id in faculty_data:
            faculty_data[faculty_id].update({
                "full_name": full_name, "department": department, "email": email
            })
            session.update({
                "full_name": full_name, "department": department, "email": email
            })
            save_json(FACULTY_FILE, faculty_data)
            flash("Profile updated successfully!", "success")
    except Exception as e:
        logger.error("faculty_update_profile error: %s", e)
        flash("Error updating profile.", "error")

    return redirect(url_for("faculty_dashboard"))


@app.route("/exam")
def exam():
    if not is_student():
        return redirect(url_for("index"))

    try:
        if "exam_question_ids" not in session:
            flash("Start exam from dashboard first.", "error")
            return redirect(url_for("dashboard"))

        questions = get_questions_by_ids(session["exam_question_ids"])
        if not questions:
            flash("Exam data missing. Please start again from the dashboard.", "error")
            return redirect(url_for("dashboard"))

        start_time = session.get("exam_start_time")
        duration = session.get("exam_duration", EXAM_DURATION_SECONDS)
        now_ts = datetime.now().timestamp()
        remaining = int(start_time + duration - now_ts)

        if remaining <= 0:
            return redirect(url_for("submit_exam"))

        practical_name = session.get("practical_name", "")
        return render_template("exam.html", questions=questions, remaining=remaining, practical_name=practical_name)
    except Exception as e:
        logger.error("exam error: %s\n%s", e, traceback.format_exc())
        flash("Error loading exam. Please try again.", "error")
        return redirect(url_for("dashboard"))


@app.route("/submit_exam", methods=["POST", "GET"])
def submit_exam():
    if not is_student():
        return redirect(url_for("index"))

    try:
        if "exam_question_ids" not in session:
            return redirect(url_for("dashboard"))

        questions = get_questions_by_ids(session["exam_question_ids"])
        if not questions:
            return redirect(url_for("dashboard"))

        submitted_answers = {}
        if request.method == "POST":
            for q in questions:
                qid = str(q["id"])
                submitted_answers[qid] = request.form.get(f"answer_{qid}")

        practical_name = (request.form.get("practical_name") or "").strip() or session.get("practical_name", "")

        total = len(questions)
        attempted = 0
        correct = 0
        detailed_answers = []

        for q in questions:
            qid = str(q["id"])
            ans = submitted_answers.get(qid)
            is_correct = bool(ans and ans == q.get("answer"))

            if ans:
                attempted += 1
                if is_correct:
                    correct += 1

            detailed_answers.append({
                "question": q.get("question", ""),
                "options": q.get("options", {}),
                "student_answer": ans if ans else "NOT ATTEMPTED",
                "correct_answer": q.get("answer", ""),
                "status": "CORRECT" if is_correct else ("WRONG" if ans else "NOT ATTEMPTED")
            })

        wrong = total - correct
        score = correct

        users = load_json(USERS_FILE)
        user = users.get(session["roll_no"])
        if not user:
            flash("User session error. Please login again.", "error")
            session.clear()
            return redirect(url_for("index"))

        dt_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        session["last_result"] = {
            "roll_no": user["roll_no"],
            "name": user.get("full_name", ""),
            "branch": user.get("branch", ""),
            "year": user.get("year", ""),
            "batch": user.get("batch", "1"),
            "email": user.get("email", ""),
            "practical_name": practical_name,
            "score": f"{score} / {total}",
            "total_questions": total,
            "attempted": attempted,
            "correct": correct,
            "wrong": wrong,
            "datetime": dt_str,
            "detailed_answers": []
        }

        os.makedirs(RESULTS_DIR, exist_ok=True)
        filename = f"Result_RollNo_{user['roll_no']}_{int(datetime.now().timestamp())}.txt"
        filepath = os.path.join(RESULTS_DIR, filename)

        lines = [
            f"Roll No: {user['roll_no']}",
            f"Name: {user.get('full_name', '')}",
            f"Branch: {user.get('branch', '')}",
            f"Year: {user.get('year', '')}",
            f"Batch: {user.get('batch', '1')}",
            f"Email: {user.get('email', '')}",
            f"Practical: {practical_name}",
            f"Score: {score} / {total}",
            f"Attempted: {attempted}",
            f"Correct: {correct}",
            f"Wrong: {wrong}",
            f"Date & Time: {dt_str}",
            "",
            "========== QUESTION WISE RESULT =========="
        ]

        for q_no, item in enumerate(detailed_answers, 1):
            lines.append("")
            lines.append(f"Q{q_no}. {item['question']}")
            options = item.get("options", {})
            for key in ["A", "B", "C", "D"]:
                lines.append(f"   {key}) {options.get(key, '')}")
            lines.append(f"Your Answer   : {item['student_answer']}")
            lines.append(f"Correct Answer: {item['correct_answer']}")
            lines.append(f"Status        : {item['status']}")
            lines.append("-" * 50)

        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("\n".join(lines))
        except Exception as e:
            logger.error("Could not write result file %s: %s", filepath, e)

        session["last_result_file"] = filename
        for key in ["exam_question_ids", "exam_start_time", "exam_duration", "practical_name"]:
            session.pop(key, None)

        return redirect(url_for("result"))
    except Exception as e:
        logger.error("submit_exam error: %s\n%s", e, traceback.format_exc())
        flash("Error submitting exam. Please try again.", "error")
        return redirect(url_for("dashboard"))


@app.route("/result")
def result():
    try:
        result_data = session.get("last_result")
        filename = session.get("last_result_file")

        if not result_data:
            flash("No result found. Please take an exam first.", "error")
            return redirect(url_for("dashboard"))

        if not result_data.get("detailed_answers") and filename:
            filepath = os.path.join(RESULTS_DIR, filename)
            if os.path.exists(filepath):
                parsed = parse_result_file(filepath)
                result_data["detailed_answers"] = parsed.get("detailed_answers", [])

        return render_template("result.html", result=result_data, filename=filename)
    except Exception as e:
        logger.error("result error: %s\n%s", e, traceback.format_exc())
        flash("Error loading result.", "error")
        return redirect(url_for("dashboard"))


@app.route("/view_result/<path:practical_name>")
def view_result(practical_name):
    if not is_student():
        return redirect(url_for("index"))

    try:
        filepath, filename = _find_result_file(session['roll_no'], practical_name)

        if not filepath:
            flash("Result not found.", "error")
            return redirect(url_for("dashboard"))

        parsed_result = parse_result_file(filepath)
        result_data = {
            "roll_no": parsed_result.get("Roll No"),
            "name": parsed_result.get("Name"),
            "branch": parsed_result.get("Branch"),
            "year": parsed_result.get("Year"),
            "batch": parsed_result.get("Batch", "1"),
            "email": parsed_result.get("Email"),
            "practical_name": parsed_result.get("Practical"),
            "score": parsed_result.get("Score"),
            "attempted": parsed_result.get("Attempted"),
            "correct": parsed_result.get("Correct"),
            "wrong": parsed_result.get("Wrong"),
            "datetime": parsed_result.get("Date & Time"),
            "detailed_answers": parsed_result.get("detailed_answers", [])
        }

        return render_template("result.html", result=result_data, filename=filename)
    except Exception as e:
        logger.error("view_result error: %s\n%s", e, traceback.format_exc())
        flash("Error loading result.", "error")
        return redirect(url_for("dashboard"))


@app.route("/download/<path:filename>")
def download(filename):
    try:
        return send_from_directory(RESULTS_DIR, filename, as_attachment=True)
    except Exception as e:
        logger.error("download error: %s", e)
        flash("File not found.", "error")
        return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("index"))


@app.route("/faculty/view_result/<roll_no>/<path:practical_name>")
def faculty_view_result(roll_no, practical_name):
    if not is_faculty():
        return redirect(url_for("index"))

    try:
        filepath, _ = _find_result_file(roll_no, practical_name)

        if not filepath:
            flash("Result not found.", "error")
            return redirect(url_for("faculty_dashboard"))

        parsed_result = parse_result_file(filepath)
        result_data = {
            "roll_no": parsed_result.get("Roll No"),
            "name": parsed_result.get("Name"),
            "branch": parsed_result.get("Branch"),
            "year": parsed_result.get("Year"),
            "batch": parsed_result.get("Batch", "1"),
            "email": parsed_result.get("Email"),
            "practical_name": parsed_result.get("Practical"),
            "score": parsed_result.get("Score"),
            "attempted": parsed_result.get("Attempted"),
            "correct": parsed_result.get("Correct"),
            "wrong": parsed_result.get("Wrong"),
            "datetime": parsed_result.get("Date & Time"),
            "detailed_answers": parsed_result.get("detailed_answers", [])
        }

        return render_template("result.html", result=result_data, filename=None, is_faculty_view=True)
    except Exception as e:
        logger.error("faculty_view_result error: %s\n%s", e, traceback.format_exc())
        flash("Error loading result.", "error")
        return redirect(url_for("faculty_dashboard"))


@app.route("/faculty/get_result_data/<roll_no>/<path:practical_name>")
def get_result_data(roll_no, practical_name):
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    try:
        filepath, _ = _find_result_file(roll_no, practical_name)
        if not filepath:
            return jsonify({"success": False, "message": "Result not found"}), 404

        parsed = parse_result_file(filepath)
        result_data = {
            "roll_no":          parsed.get("Roll No"),
            "name":             parsed.get("Name"),
            "branch":           parsed.get("Branch"),
            "year":             parsed.get("Year"),
            "batch":            parsed.get("Batch", "1"),
            "email":            parsed.get("Email"),
            "practical_name":   parsed.get("Practical"),
            "score":            parsed.get("Score"),
            "attempted":        parsed.get("Attempted"),
            "correct":          parsed.get("Correct"),
            "wrong":            parsed.get("Wrong"),
            "datetime":         parsed.get("Date & Time"),
            "detailed_answers": parsed.get("detailed_answers", [])
        }
        return jsonify({"success": True, "result": result_data})
    except Exception as e:
        logger.error("get_result_data error: %s", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/faculty/get_result_txt/<roll_no>/<path:practical_name>")
def get_result_txt(roll_no, practical_name):
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401

    try:
        filepath, filename = _find_result_file(roll_no, practical_name)
        if not filepath:
            return jsonify({"success": False, "message": "File not found"}), 404

        file_content = read_file_safe(filepath)
        if file_content is None:
            return jsonify({"success": False, "message": "Could not read file"}), 500

        return jsonify({"success": True, "content": file_content, "filename": filename})
    except Exception as e:
        logger.error("get_result_txt error: %s", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/api/get_questions")
def get_questions():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        practical_name = request.args.get("practical", "").strip()
        if not practical_name:
            return jsonify({"success": False, "message": "Practical name required"}), 400
        questions = load_json(QUESTIONS_FILE)
        practical_questions = [q for q in questions if q.get("practical") == practical_name]
        return jsonify({"success": True, "questions": practical_questions, "count": len(practical_questions)})
    except Exception as e:
        logger.error("get_questions error: %s", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/api/add_question", methods=["POST"])
def add_question():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json(silent=True) or {}
        practical_name = data.get("practical", "").strip()
        question_text = data.get("question", "").strip()
        options = data.get("options", {})
        answer = data.get("answer", "").strip()

        if not practical_name:
            return jsonify({"success": False, "message": "Practical name required"}), 400
        if not question_text:
            return jsonify({"success": False, "message": "Question text required"}), 400
        if not all(str(options.get(k, "")).strip() for k in ["A", "B", "C", "D"]):
            return jsonify({"success": False, "message": "All 4 options are required"}), 400
        if answer not in ["A", "B", "C", "D"]:
            return jsonify({"success": False, "message": "Correct answer must be A, B, C or D"}), 400

        questions = load_json(QUESTIONS_FILE)
        existing = [q for q in questions if q.get("practical") == practical_name]
        if len(existing) >= 20:
            return jsonify({"success": False, "message": "Maximum 20 questions allowed per practical"}), 400

        try:
            new_id = max((int(q.get("id", 0)) for q in questions), default=0) + 1
        except (ValueError, TypeError):
            new_id = len(questions) + 1

        new_question = {
            "id": new_id,
            "practical": practical_name,
            "question": question_text,
            "options": {
                "A": str(options.get("A", "")).strip(),
                "B": str(options.get("B", "")).strip(),
                "C": str(options.get("C", "")).strip(),
                "D": str(options.get("D", "")).strip()
            },
            "answer": answer
        }
        questions.append(new_question)
        save_json(QUESTIONS_FILE, questions)
        return jsonify({"success": True, "question": new_question, "total": len(existing) + 1}), 200
    except Exception as e:
        logger.error("add_question error: %s\n%s", e, traceback.format_exc())
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


@app.route("/api/delete_question", methods=["POST"])
def delete_question():
    if not is_faculty():
        return jsonify({"success": False, "message": "Unauthorized"}), 401
    try:
        data = request.get_json(silent=True) or {}
        question_id = data.get("id")
        if question_id is None:
            return jsonify({"success": False, "message": "Question ID required"}), 400
        questions = load_json(QUESTIONS_FILE)
        original_len = len(questions)
        questions = [q for q in questions if q.get("id") != question_id]
        if len(questions) == original_len:
            return jsonify({"success": False, "message": "Question not found"}), 404
        save_json(QUESTIONS_FILE, questions)
        return jsonify({"success": True}), 200
    except Exception as e:
        logger.error("delete_question error: %s", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500


if __name__ == "__main__":
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)
    app.run(host='0.0.0.0', port=5000, debug=True)
